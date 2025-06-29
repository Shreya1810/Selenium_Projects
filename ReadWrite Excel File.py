import openpyxl

book = openpyxl.load_workbook("C:\\Users\\SinhaShreya\\Downloads\\Important files\\Python module.xlsx")
sheet = book.active
# empty dict
Dict= {}
focus_cell = sheet.cell(row=1,column=2)
print(focus_cell.value)

# write in that sheet
sheet.cell(row=2,column=2).value = "Shreya"
#print(sheet.cell(row=2,column=2).value)

# count total rows and columns present in this sheet
#print(sheet.max_row)
#print(sheet.max_column)
#print(sheet['A7'].value)

# print all the data using loop
for r in range(1,sheet.max_row+1): # to get rows
    if sheet.cell(row=r, column=1).value == "Testcase 2":
        for c in range(2,sheet.max_column+1): # to get columns
            # Dict["lastname"] = "Sinha" --> example
            Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r,column=c).value
print(Dict)
